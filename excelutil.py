from openpyxl import load_workbook


def table_to_dictlist (file: str, sheet: str, start_coordinates: tuple[int, int], end_coordinates: tuple[int, int]) -> list[dict]:
    """ Reads an area of an excel sheet and returns the information as list of dictionaries. 
        Assumes the first row to be a headline.

    Args:
        file (str): Path to the excel file.
        sheet (str): Name of the sheet in the excel file.
        start_coordinates (tuple): Top-left coordinates of the area to be read. (E.g. B3 = (2,3))
        end_coordinates (tuple): Bottem-right coordinates of the area to be read. (E.g. F5 = (6,2))

    Returns:
        list[dict]: List of dictionaries where each list represents a row from the excel area and each dictionary 
            represents a value in this row with the corresponding headline-value as key
    """
    
    # Load Worksheet
    sheet = load_workbook(filename=file)[sheet]
    
    # Get dictionary keys from headline
    headline = sheet.iter_rows(min_row = start_coordinates[1], max_row = start_coordinates[1], min_col = start_coordinates[0], max_col = end_coordinates[0], values_only=True)
    keys = next(headline)

    # Read values
    l = []
    for row in sheet.iter_rows(min_row = start_coordinates[1] + 1, max_row = end_coordinates[1], min_col = start_coordinates[0], max_col = end_coordinates[0], values_only=True):
        d = {}
        for key, value in zip (keys, row):
            d[key] = value

        l.append(d)

    return l